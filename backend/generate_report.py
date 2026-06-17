import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

data = [
    ["01", "Đăng ký tài khoản thành công", "Hệ thống tạo tài khoản thành công, hiển thị thông báo 'Register success'", "Tạo tài khoản không thành công (Mong đợi: Thành công)", "FAIL"],
    ["02", "Đăng ký thất bại khi nhập sai Password", "Hệ thống hiển thị lỗi 'Invalid password'", "Hệ thống chặn đăng ký và báo lỗi", "PASS"],
    ["03", "Đăng ký thất bại khi để trống Password", "Hệ thống hiển thị lỗi 'Password không được để trống'", "Alert: Invalid password: Must be 6-50 characters.", "PASS"],
    ["04", "Đăng ký thất bại khi nhập sai Phone", "Hệ thống hiển thị lỗi yêu cầu nhập đúng định dạng Phone", "Alert: Invalid phone: Must be 10-11 digits.", "PASS"],
    ["05", "Đăng ký thất bại khi để trống Phone", "Hệ thống yêu cầu nhập trường Phone", "Alert: Invalid phone: Must be 10-11 digits.", "PASS"],
    ["06", "Đăng ký thất bại khi nhập sai định dạng Email", "Hệ thống hiển thị lỗi Email không đúng định dạng", "Alert: Invalid email format.", "PASS"],
    ["07", "Đăng ký thất bại khi để trống Email", "Hệ thống yêu cầu nhập Email", "Alert: Invalid email format.", "PASS"],
    ["08", "Đăng ký thất bại khi nhập FullName không hợp lệ", "Hệ thống hiển thị lỗi FullName không hợp lệ", "Hệ thống chặn đăng ký và báo lỗi", "PASS"],
    ["09", "Đăng ký thất bại khi để trống FullName", "Hệ thống yêu cầu nhập FullName", "Alert: Invalid name: Must be 2-50 characters.", "PASS"]
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Report"

# Thêm tiêu đề lớn
ws.append(["BÁO CÁO KẾT QUẢ KIỂM THỬ TỰ ĐỘNG: ĐĂNG KÝ (REGISTER BQD)"])
ws.merge_cells('A1:E1')
title_cell = ws['A1']
title_cell.font = Font(name='Arial', size=14, bold=True, color="1F497D")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Dòng trống
ws.append([])

# Headers
headers = ["ID", "Tên Test Case (Title)", "Kết Quả Mong Đợi (Expected Result)", "Kết Quả Thực Tế (Actual Result)", "Trạng Thái"]
ws.append(headers)

header_row = 3
header_font = Font(name='Arial', bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")
alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for col_num, cell in enumerate(ws[header_row], 1):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment_center
    cell.border = thin_border

# Data Rows
pass_font = Font(name='Arial', bold=True, color="008000") # Green
fail_font = Font(name='Arial', bold=True, color="FF0000") # Red
normal_font = Font(name='Arial')
alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

for row_data in data:
    ws.append(row_data)
    row_idx = ws.max_row
    
    for col_idx, cell in enumerate(ws[row_idx], 1):
        cell.border = thin_border
        cell.font = normal_font
        if col_idx == 1 or col_idx == 5:
            cell.alignment = alignment_center
        else:
            cell.alignment = alignment_left
            
        # Format Pass/Fail
        if col_idx == 5:
            if cell.value == "PASS":
                cell.font = pass_font
            elif cell.value == "FAIL":
                cell.font = fail_font

# Chỉnh độ rộng cột
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 50
ws.column_dimensions['E'].width = 15

# Lưu file
report_path = r"d:\ThuHang\ThuHang\doc\Bao_Cao_Test_Register_BQD.xlsx"
wb.save(report_path)
print("Successfully generated Excel report at:", report_path)
